"""
旧格式解析工具（.doc / .xls）

目标：满足 REQ-NF-007 的隔离/超时/清理要求，为多入口复用。

实现策略：
- 在隔离临时目录内写入上传文件副本
- .doc 通过 aspose-words-foss 提取文本（带超时）
- .xls 通过 xlrd 直接提取表格内容
- 解析成功/失败均清理临时目录
"""

from __future__ import annotations

import importlib
import logging
import multiprocessing
import os
import queue
import re
import tempfile
from typing import Any, Dict, List, Optional

from backend.config.config import settings

logger = logging.getLogger(__name__)

OLE_COMPOUND_FILE_SIGNATURE = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def _get_timeout_seconds(override: Optional[int] = None) -> int:
    if override is not None:
        try:
            return max(1, int(override))
        except Exception:
            return 60
    try:
        value = int(os.getenv("OLD_FORMAT_PARSE_TIMEOUT_SECONDS", "60"))
    except Exception:
        value = 60
    return max(1, min(value, 600))


def _safe_stem(filename: str, default: str = "upload") -> str:
    text = str(filename or "").strip() or default
    text = os.path.basename(text.replace("\\", "/")).replace("\x00", "").strip()
    stem = os.path.splitext(text)[0].strip() or default
    stem = re.sub(r"[^A-Za-z0-9._-]+", "_", stem).strip("._")
    return stem or default


def _old_format_tmp_root() -> str:
    root = os.path.join(str(settings.UPLOAD_DIR or "uploads"), "_old_format_parse")
    os.makedirs(root, exist_ok=True)
    return root


def _doc_file_to_text(input_path: str) -> str:
    errors: List[str] = []
    aw = None
    for module_name in ("aspose.words_foss", "aspose.words"):
        try:
            aw = importlib.import_module(module_name)
            break
        except Exception as exc:
            errors.append(f"{module_name}: {exc}")

    if aw is None:
        detail = "; ".join(errors)
        raise RuntimeError(f"旧格式解析工具不可用：aspose-words 未安装或不可导入 ({detail})")

    try:
        return (aw.Document(input_path).get_text() or "").strip()
    except Exception as exc:
        raise RuntimeError(f"旧格式解析失败: {str(exc)[:200]}") from exc


def _doc_worker(input_path: str, result_queue: multiprocessing.Queue) -> None:
    try:
        result_queue.put(("ok", _doc_file_to_text(input_path)))
    except BaseException as exc:
        result_queue.put(("error", str(exc)))


def _run_doc_file_to_text(input_path: str, timeout_seconds: int) -> str:
    start_method = "fork" if "fork" in multiprocessing.get_all_start_methods() else None
    context = multiprocessing.get_context(start_method)
    result_queue = context.Queue(maxsize=1)
    process = context.Process(target=_doc_worker, args=(input_path, result_queue))

    process.start()
    process.join(timeout_seconds)

    if process.is_alive():
        process.terminate()
        process.join(5)
        raise TimeoutError("旧格式解析超时")

    try:
        status, payload = result_queue.get_nowait()
    except queue.Empty:
        raise RuntimeError(f"旧格式解析失败: 子进程退出码 {process.exitcode}")

    if status == "ok":
        return str(payload or "").strip()
    raise RuntimeError(str(payload or "旧格式解析失败"))


def doc_bytes_to_text(file_content: bytes, filename: str, *, timeout_seconds: Optional[int] = None) -> str:
    timeout = _get_timeout_seconds(timeout_seconds)
    stem = _safe_stem(filename, default="document")
    content = file_content or b""
    if not content.startswith(OLE_COMPOUND_FILE_SIGNATURE):
        raise RuntimeError("旧格式解析失败: 无效的DOC文件格式")

    with tempfile.TemporaryDirectory(prefix="doc_", dir=_old_format_tmp_root()) as tmpdir:
        input_path = os.path.join(tmpdir, f"{stem}.doc")
        with open(input_path, "wb") as f:
            f.write(content)

        text = (_run_doc_file_to_text(input_path, timeout) or "").strip()
        if not text:
            raise RuntimeError("旧格式解析失败: 未提取到文本")
        return text


def xlsx_bytes_to_sheet_rows(xlsx_bytes: bytes) -> Dict[str, List[List[Any]]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("缺少Excel解析依赖 openpyxl") from exc

    from io import BytesIO

    wb = load_workbook(BytesIO(xlsx_bytes), data_only=True)
    data: Dict[str, List[List[Any]]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_rows: List[List[Any]] = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None and str(cell).strip() != "" for cell in row):
                sheet_rows.append(list(row))
        data[sheet_name] = sheet_rows
    return data


def _xls_bytes_to_sheet_rows_with_xlrd(file_content: bytes) -> Dict[str, List[List[Any]]]:
    try:
        import xlrd
    except Exception as exc:
        raise RuntimeError("旧格式解析工具不可用：xlrd未安装") from exc

    try:
        workbook = xlrd.open_workbook(file_contents=file_content)
    except Exception as exc:
        raise RuntimeError(f"旧格式解析失败: {str(exc)[:200]}") from exc

    data: Dict[str, List[List[Any]]] = {}
    for sheet in workbook.sheets():
        rows: List[List[Any]] = []
        for row_idx in range(sheet.nrows):
            row_values = sheet.row_values(row_idx)
            if any(cell is not None and str(cell).strip() != "" for cell in row_values):
                rows.append(list(row_values))
        data[sheet.name] = rows
    return data


def xls_bytes_to_sheet_rows(file_content: bytes, filename: str, *, timeout_seconds: Optional[int] = None) -> Dict[str, List[List[Any]]]:
    return _xls_bytes_to_sheet_rows_with_xlrd(file_content)


def sheet_rows_to_text(sheet_rows: Dict[str, List[List[Any]]], *, max_lines: int = 2000) -> str:
    lines: List[str] = []
    for sheet_name, rows in (sheet_rows or {}).items():
        if sheet_name:
            lines.append(f"[Sheet] {sheet_name}")
        for row in rows or []:
            if not isinstance(row, list):
                continue
            line = " | ".join(str(cell).strip() for cell in row if cell is not None and str(cell).strip())
            if line:
                lines.append(line)
            if len(lines) >= max_lines:
                break
        if len(lines) >= max_lines:
            break
    return "\n".join(lines).strip()


def xls_bytes_to_text(file_content: bytes, filename: str, *, timeout_seconds: Optional[int] = None) -> str:
    sheet_rows = xls_bytes_to_sheet_rows(file_content, filename, timeout_seconds=timeout_seconds)
    return sheet_rows_to_text(sheet_rows)
