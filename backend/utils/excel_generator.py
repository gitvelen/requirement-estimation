"""
Excel报告生成模块
生成工作量评估Excel报告
"""
import os
import logging
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from backend.config.config import settings

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """Excel报告生成器"""

    def __init__(self):
        """初始化生成器"""
        self.report_dir = settings.REPORT_DIR
        os.makedirs(self.report_dir, exist_ok=True)

        # 定义样式
        self.header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.cell_alignment = Alignment(vertical='top', wrap_text=True)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        logger.info("Excel生成器初始化完成")

    def generate_report(
        self,
        task_id: str,
        requirement_name: str,
        systems_data: Dict[str, List[Dict[str, Any]]],
        expert_estimates: Optional[Dict[str, List[float]]] = None
    ) -> str:
        """
        生成Excel报告

        Args:
            task_id: 任务ID
            requirement_name: 需求名称
            systems_data: 各系统的功能点数据
                格式: {
                    "系统名": [
                        {
                            "序号": "1.1",
                            "功能模块": "模块A",
                            "功能点": "功能点1",
                            "业务描述": "...",
                            "输入": "...",
                            "输出": "...",
                            "依赖": "...",
                            "预估人天": 5,
                            "复杂度": "高",
                            "备注": "..."
                        },
                        ...
                    ],
                    ...
                }
            expert_estimates: 专家估算数据（可选）
                格式: {
                    "系统名": [专家1估值, 专家2估值, ..., 专家5估值],
                    ...
                }

        Returns:
            str: 生成的Excel文件路径
        """
        try:
            import time
            start_time = time.time()

            logger.info(f"[Excel生成] 开始生成报告")
            logger.info(f"[任务信息] ID: {task_id}, 需求: {requirement_name}")

            # 创建工作簿
            wb = Workbook()
            wb.remove(wb.active)

            # 创建汇总sheet
            logger.info(f"[处理中] 创建汇总统计sheet...")
            self._create_summary_sheet(wb, requirement_name, systems_data, expert_estimates)

            # 为每个系统创建一个sheet
            logger.info(f"[处理中] 创建 {len(systems_data)} 个系统sheet...")
            for idx, (system_name, features) in enumerate(systems_data.items(), 1):
                logger.info(f"  [{idx}/{len(systems_data)}] 创建系统sheet: {system_name} ({len(features)} 个功能点)")
                self._create_system_sheet(wb, system_name, features, expert_estimates)

            # 生成文件名
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{task_id}_{requirement_name[:20]}_{timestamp}.xlsx"
            file_path = os.path.join(self.report_dir, filename)

            # 保存文件
            wb.save(file_path)

            # 计算文件大小
            file_size = os.path.getsize(file_path) / 1024  # KB
            elapsed_time = time.time() - start_time

            logger.info(f"[Excel生成] 报告生成成功")
            logger.info(f"[文件信息]")
            logger.info(f"  - 路径: {file_path}")
            logger.info(f"  - 大小: {file_size:.1f} KB")
            logger.info(f"  - 系统数: {len(systems_data)}")
            logger.info(f"  - 耗时: {elapsed_time:.2f} 秒")

            return file_path

        except Exception as e:
            logger.error(f"[Excel生成] 报告生成失败: {str(e)}")
            raise

    def _create_summary_sheet(
        self,
        wb: Workbook,
        requirement_name: str,
        systems_data: Dict[str, List[Dict[str, Any]]],
        expert_estimates: Optional[Dict[str, List[float]]] = None
    ):
        """
        创建汇总sheet

        Args:
            wb: 工作簿对象
            requirement_name: 需求名称
            systems_data: 各系统的功能点数据
            expert_estimates: 专家估算数据
        """
        ws = wb.create_sheet(title="汇总统计", index=0)

        # 写入标题
        ws['A1'] = '业务需求工作量评估报告'
        ws['A1'].font = Font(name='微软雅黑', size=16, bold=True)
        ws.merge_cells('A1:I1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # 写入基本信息
        ws['A2'] = f'需求名称: {requirement_name}'
        ws['A3'] = f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws['A4'] = f'评估方法: Delphi专家评估法（1-3轮评分）+ COSMIC功能点分析'
        ws['A5'] = f'说明：请在各系统sheet手工填写专家评估人天，系统自动取最后一轮有值的结果进行汇总（第3轮>第2轮>第1轮）'

        # 构建系统数据列表，记录每个系统的小计行位置
        system_info_list = []

        for system_name, features in systems_data.items():
            high = 0
            mid = 0
            low = 0
            modules = set()

            for feature in features:
                complexity = feature.get("复杂度", "")
                if complexity == "高":
                    high += 1
                elif complexity == "中":
                    mid += 1
                elif complexity == "低":
                    low += 1
                modules.add(feature.get("功能模块", ""))

            # 记录系统信息（不直接写入工作量，使用公式）
            system_info_list.append({
                "name": system_name,
                "features_count": len(features),
                "high": high,
                "mid": mid,
                "low": low,
                "modules": sorted(modules),
                "sheet_name": system_name[:31]  # Excel sheet名称限制
            })

        # 写入汇总数据行（使用公式引用各系统sheet）
        start_row = 6

        # 首先写入表头（第6行）
        headers = ["系统名称", "功能点数", "最终专家评估均值(人天)", "高复杂度", "中复杂度", "低复杂度", "工作量占比", "主要功能模块"]
        for col_idx, header_text in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=header_text)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border

        current_row = start_row + 1  # 跳过表头

        for idx, sys_info in enumerate(system_info_list):
            sheet_name = sys_info["sheet_name"]
            feature_count = sys_info["features_count"]
            # 小计行号 = 表头行(1) + 数据行数 + 1
            subtotal_row = feature_count + 2

            # 构建公式：智能判断取哪一轮的均值
            # 优先取第3轮(V列)，如果为空则取第2轮(R列)，如果还空则取第1轮(N列)
            # 使用IF嵌套：=IF(V小计行<>"", V小计行, IF(R小计行<>"", R小计行, N小计行))
            workload_formula = (
                f"=IF('{sheet_name}'!V{subtotal_row}<>\"\", '{sheet_name}'!V{subtotal_row}, "
                f"IF('{sheet_name}'!R{subtotal_row}<>\"\", '{sheet_name}'!R{subtotal_row}, "
                f"'{sheet_name}'!N{subtotal_row}))"
            )

            # 计算工作量占比
            ratio_formula = f"=C{current_row}/SUM(C${start_row + 1}:C${start_row + len(system_info_list)})"

            ws.cell(row=current_row, column=1, value=sys_info["name"])
            ws.cell(row=current_row, column=2, value=feature_count)
            ws.cell(row=current_row, column=3, value=workload_formula)  # 使用智能公式
            ws.cell(row=current_row, column=4, value=sys_info["high"])
            ws.cell(row=current_row, column=5, value=sys_info["mid"])
            ws.cell(row=current_row, column=6, value=sys_info["low"])
            ws.cell(row=current_row, column=7, value=ratio_formula)
            ws.cell(row=current_row, column=8, value="、".join(sys_info["modules"]))

            current_row += 1

        # 合计行（使用SUM公式）
        total_row = current_row
        ws.cell(row=total_row, column=1, value="合计")
        ws.cell(row=total_row, column=2, value=f"=SUM(B{start_row + 1}:B{total_row - 1})")
        ws.cell(row=total_row, column=3, value=f"=SUM(C{start_row + 1}:C{total_row - 1})")
        ws.cell(row=total_row, column=4, value=f"=SUM(D{start_row + 1}:D{total_row - 1})")
        ws.cell(row=total_row, column=5, value=f"=SUM(E{start_row + 1}:E{total_row - 1})")
        ws.cell(row=total_row, column=6, value=f"=SUM(F{start_row + 1}:F{total_row - 1})")
        ws.cell(row=total_row, column=7, value="100%")
        ws.cell(row=total_row, column=8, value="-")

        # 设置数据行和合计行样式
        for row_idx in range(start_row + 1, total_row + 1):
            for col_idx in range(1, 9):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                cell.border = self.border
                if row_idx == total_row:  # 合计行
                    cell.font = Font(name='微软雅黑', size=11, bold=True)

        # 调整列宽
        column_widths = [20, 12, 15, 12, 12, 12, 12, 50]
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def _create_system_sheet(
        self,
        wb: Workbook,
        system_name: str,
        features: List[Dict[str, Any]],
        expert_estimates: Optional[Dict[str, List[float]]] = None
    ):
        """
        创建系统sheet

        Args:
            wb: 工作簿对象
            system_name: 系统名称
            features: 功能点列表
            expert_estimates: 专家估算数据
        """
        # sheet名称最多31字符
        sheet_name = system_name[:31]
        ws = wb.create_sheet(title=sheet_name)

        # 准备表头 - 基础信息列
        headers = ["序号", "功能模块", "功能点", "业务描述", "输入", "输出", "依赖项", "复杂度等级", "备注", "预估人天参考"]

        # 添加3轮专家评估列（每轮：专家1、专家2、专家3、均值）
        for round_num in range(1, 4):  # 3轮
            headers.extend([
                f"第{round_num}轮专家1",
                f"第{round_num}轮专家2",
                f"第{round_num}轮专家3",
                f"第{round_num}轮均值"
            ])

        # 写入表头和数据
        rows = [headers]

        for idx, feature in enumerate(features, start=1):
            feature_name = feature.get("功能点", "")

            # 基础信息 + 预估人天参考值
            row = [
                feature.get("序号", f"{idx}"),
                feature.get("功能模块", ""),
                feature_name,
                feature.get("业务描述", ""),
                feature.get("输入", ""),
                feature.get("输出", ""),
                feature.get("依赖", ""),
                feature.get("复杂度", ""),
                feature.get("备注", ""),
                float(feature.get("预估人天", 2.5))  # 预估人天参考
            ]

            # 为3轮专家评估添加空值（后续在写入时处理均值公式）
            # 每轮：专家1、专家2、专家3、均值
            for round_num in range(3):  # 3轮
                row.extend([None, None, None, None])  # 专家1、专家2、专家3、均值

            rows.append(row)

        # 写入小计行
        if features:
            # 小计行：基础列 + 预估人天参考 + 3轮专家评估
            last_row_num = len(features) + 1  # 数据行数 + 表头行
            row_idx = last_row_num  # Excel中实际的行号

            # 前8列为空，第9列（备注列）是"小计"
            subtotal_row = ["", "", "", "", "", "", "", "", "小计"]

            # 预估人天参考小计（J列）
            subtotal_row.append(f"=SUM(J2:J{row_idx})")

            # 为3轮专家评估添加SUM公式
            # 第1轮：K、L、M、N列
            subtotal_row.extend([
                f"=SUM(K2:K{row_idx})",  # 第1轮专家1
                f"=SUM(L2:L{row_idx})",  # 第1轮专家2
                f"=SUM(M2:M{row_idx})",  # 第1轮专家3
                f"=SUM(N2:N{row_idx})",  # 第1轮均值
            ])

            # 第2轮：O、P、Q、R列
            subtotal_row.extend([
                f"=SUM(O2:O{row_idx})",  # 第2轮专家1
                f"=SUM(P2:P{row_idx})",  # 第2轮专家2
                f"=SUM(Q2:Q{row_idx})",  # 第2轮专家3
                f"=SUM(R2:R{row_idx})",  # 第2轮均值
            ])

            # 第3轮：S、T、U、V列
            subtotal_row.extend([
                f"=SUM(S2:S{row_idx})",  # 第3轮专家1
                f"=SUM(T2:T{row_idx})",  # 第3轮专家2
                f"=SUM(U2:U{row_idx})",  # 第3轮专家3
                f"=SUM(V2:V{row_idx})",  # 第3轮均值
            ])

            rows.append(subtotal_row)

        # 写入数据
        for row_idx, row_data in enumerate(rows, start=1):
            for col_idx, cell_value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)

                if row_idx == 1:  # 表头
                    cell.value = cell_value
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = self.header_alignment
                    cell.border = self.border
                elif row_idx == len(rows):  # 小计行
                    cell.value = cell_value
                    cell.font = Font(name='微软雅黑', size=11, bold=True)
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.border = self.border
                else:
                    # 数据行
                    if cell_value is not None:
                        cell.value = cell_value
                    cell.alignment = self.cell_alignment
                    cell.border = self.border

                    # 为均值列写入Excel公式
                    # 第1轮均值：第14列（N列）
                    if col_idx == 14:
                        cell.value = f"=AVERAGE(K{row_idx}:M{row_idx})"
                    # 第2轮均值：第18列（R列）
                    elif col_idx == 18:
                        cell.value = f"=AVERAGE(O{row_idx}:Q{row_idx})"
                    # 第3轮均值：第22列（V列）
                    elif col_idx == 22:
                        cell.value = f"=AVERAGE(S{row_idx}:U{row_idx})"

        # 调整列宽
        # 前10列：基础信息；后12列：3轮专家评估（每轮4列）
        column_widths = [
            8,   # 序号
            15,  # 功能模块
            20,  # 功能点
            40,  # 业务描述
            20,  # 输入
            20,  # 输出
            15,  # 依赖项
            10,  # 复杂度等级
            30,  # 备注
            12,  # 预估人天参考
            # 第1轮专家评估
            10, 10, 10, 12,  # 专家1、专家2、专家3、均值
            # 第2轮专家评估
            10, 10, 10, 12,  # 专家1、专家2、专家3、均值
            # 第3轮专家评估
            10, 10, 10, 12,  # 专家1、专家2、专家3、均值
        ]
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[1].height = 30

        # 冻结窗格：冻结前10列（A-J）和第一行
        ws.freeze_panes = "K2"


# 全局生成器实例
excel_generator = ExcelGenerator()
