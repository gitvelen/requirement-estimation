"""
系统清单配置 API（XLSX 批量导入）
支持从单一系统清单 Excel 文件中导入铺底数据。
"""

from __future__ import annotations

import logging
import os
import re
import uuid
import zipfile
from datetime import date, datetime
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, Depends, File, Header, HTTPException, Request, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from backend.api import system_routes
from backend.api.auth import require_admin_api_key
from backend.api.error_utils import ApiError, build_error_response
from backend.config.config import settings
from backend.service.system_catalog_profile_initializer import get_system_catalog_profile_initializer

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/v1/system-list", tags=["系统清单配置"])

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
TEMPLATE_PATH = os.path.join(BASE_DIR, "data", "syslist-template.xlsx")


class SystemListImportConfirmRequest(BaseModel):
    mode: str = Field("replace", description="replace=覆盖导入；upsert=增量导入")
    systems: List[Dict[str, Any]] = Field(default_factory=list)


async def require_system_list_admin(
    x_api_key: Optional[str] = Header(None),
    authorization: Optional[str] = Header(None),
) -> None:
    try:
        require_admin_api_key(x_api_key=x_api_key, authorization=authorization)
    except HTTPException as exc:
        status_code = 403
        if int(exc.status_code or 0) >= 500:
            status_code = int(exc.status_code)
        raise ApiError(
            error_code="AUTH_001",
            message="权限不足",
            status_code=status_code,
            details={"reason": str(exc.detail)},
        ) from exc


def _to_api_error(exc: Exception, *, default_error_code: str = "CATALOG_001") -> ApiError:
    if isinstance(exc, ApiError):
        return exc
    if isinstance(exc, HTTPException):
        status_code = int(exc.status_code or 400)
        error_code = default_error_code
        if status_code == 403:
            error_code = "AUTH_001"
        return ApiError(
            error_code=error_code,
            message=str(exc.detail or "系统清单处理失败"),
            status_code=status_code,
        )
    return ApiError(
        error_code=default_error_code,
        message="系统清单处理失败",
        status_code=400,
        details={"reason": str(exc)},
    )


def _normalize_system_extra(extra: Any) -> Dict[str, Any]:
    if not isinstance(extra, dict):
        return {}

    normalized_extra: Dict[str, Any] = {}
    for raw_key, raw_value in extra.items():
        canonical_key = system_routes.resolve_owner_extra_key(str(raw_key or ""))
        key = canonical_key or str(raw_key or "").strip()
        if not key:
            continue
        if raw_value is None:
            normalized_extra[key] = ""
        elif isinstance(raw_value, (list, dict)):
            normalized_extra[key] = raw_value
        else:
            normalized_extra[key] = str(raw_value).strip()

    normalized_extra, _ = system_routes.normalize_system_owner_extra_fields(normalized_extra)
    return normalized_extra


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", "").replace("\r", "").strip()


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def _cell_to_int(value: Any) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return None


def _sanitize_xlsx_for_openpyxl(content: bytes) -> bytes:
    """
    当前环境的 openpyxl 无法解析部分带 dataValidation.id 的工作簿。
    这里移除 worksheet xml 中的 dataValidations 片段，保持导入可用。
    """

    in_mem = BytesIO(content)
    out_mem = BytesIO()
    try:
        with zipfile.ZipFile(in_mem, "r") as zin, zipfile.ZipFile(out_mem, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename.startswith("xl/worksheets/") and item.filename.endswith(".xml"):
                    try:
                        text = data.decode("utf-8")
                        text = re.sub(r"<dataValidations[^>]*/>", "", text, flags=re.DOTALL)
                        text = re.sub(r"<dataValidations[^>]*>[\s\S]*?</dataValidations>", "", text, flags=re.DOTALL)
                        data = text.encode("utf-8")
                    except UnicodeDecodeError:
                        pass
                zout.writestr(item, data)
        return out_mem.getvalue()
    except zipfile.BadZipFile as exc:
        raise ApiError(
            error_code="CATALOG_001",
            message="系统清单文件不合法",
            status_code=400,
            details={"reason": "无效的XLSX文件"},
        ) from exc


def _load_workbook(content: bytes):
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise ApiError(
            error_code="CATALOG_001",
            message="系统清单文件不合法",
            status_code=500,
            details={"reason": "缺少Excel解析依赖 openpyxl"},
        ) from exc

    sanitized = _sanitize_xlsx_for_openpyxl(content)
    try:
        return load_workbook(BytesIO(sanitized), data_only=True)
    except Exception as exc:
        logger.exception("解析XLSX失败")
        raise ApiError(
            error_code="CATALOG_001",
            message="系统清单文件不合法",
            status_code=400,
            details={"reason": f"解析XLSX失败: {str(exc)}"},
        ) from exc


def _find_header_row(ws, header_aliases: Dict[str, List[str]]) -> Tuple[int, Dict[str, int]]:
    required_keys = [key for key in header_aliases if key in {"name", "abbreviation"}]
    for row_index, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        headers = [_normalize_header(cell) for cell in row]
        raw_map: Dict[str, int] = {}
        for col_index, header in enumerate(headers):
            if header and header not in raw_map:
                raw_map[header] = col_index

        canonical_map: Dict[str, int] = {}
        for canonical, aliases in header_aliases.items():
            for alias in aliases:
                if alias in raw_map:
                    canonical_map[canonical] = raw_map[alias]
                    break

        if all(key in canonical_map for key in required_keys):
            return row_index, canonical_map

    raise ApiError(
        error_code="CATALOG_002",
        message="系统清单模板不匹配",
        status_code=400,
        details={"reason": "模板表头无法识别，请使用最新系统清单模板"},
    )


def _detect_system_sheet(wb):
    aliases = {
        "name": ["系统名称", "系统名", "系统", "name", "system_name"],
        "abbreviation": ["英文简称", "系统简称", "英文缩写", "abbreviation", "abbr"],
    }
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try:
            _find_header_row(ws, aliases)
            return ws
        except ApiError:
            continue

    raise ApiError(
        error_code="CATALOG_002",
        message="系统清单模板不匹配",
        status_code=400,
        details={"reason": "未找到可识别的系统清单工作表"},
    )


def _validate_systems(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    for item in rows:
        item.setdefault("errors", [])
        if not str(item.get("name") or "").strip():
            item["errors"].append("系统名称不能为空")
        if not str(item.get("abbreviation") or "").strip():
            item["errors"].append("英文简称不能为空")
    return rows


def _parse_systems_sheet(ws) -> List[Dict[str, Any]]:
    header_row, header_map = _find_header_row(
        ws,
        {
            "name": ["系统名称", "系统名", "系统", "name", "system_name"],
            "abbreviation": ["英文简称", "系统简称", "英文缩写", "abbreviation", "abbr"],
            "status": ["状态", "系统状态", "status"],
        },
    )
    name_idx = header_map["name"]
    abbr_idx = header_map["abbreviation"]
    status_idx = header_map.get("status")

    header_values = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True), [])
    normalized_headers = [_normalize_header(cell) for cell in header_values]
    seen_headers: Dict[str, int] = {}
    index_to_header: Dict[int, str] = {}
    for col_index, raw_header in enumerate(normalized_headers):
        if not raw_header:
            continue
        if raw_header in seen_headers:
            seen_headers[raw_header] += 1
            header_key = f"{raw_header}#{seen_headers[raw_header]}"
        else:
            seen_headers[raw_header] = 1
            header_key = raw_header
        index_to_header[col_index] = header_key

    core_indices = {name_idx, abbr_idx}
    if status_idx is not None:
        core_indices.add(status_idx)

    results: List[Dict[str, Any]] = []
    for row_number, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if not any(_cell_to_text(cell) for cell in row):
            continue

        name = _cell_to_text(row[name_idx] if name_idx < len(row) else None)
        abbr = _cell_to_text(row[abbr_idx] if abbr_idx < len(row) else None)
        status = _cell_to_text(row[status_idx] if status_idx is not None and status_idx < len(row) else None)

        extra: Dict[str, Any] = {}
        for col_index, header_key in index_to_header.items():
            if col_index in core_indices:
                continue
            extra[header_key] = _cell_to_text(row[col_index] if col_index < len(row) else None)

        results.append(
            {
                "row_number": row_number,
                "name": name,
                "abbreviation": abbr,
                "status": status,
                "extra": _normalize_system_extra(extra),
                "errors": [],
            }
        )
    return _validate_systems(results)


def _reload_system_identification_cache() -> None:
    try:
        from backend.agent.system_identification_agent import get_system_identification_agent

        system_identification_agent = get_system_identification_agent()
        system_identification_agent.system_list = system_identification_agent._load_system_list()
    except Exception:
        logger.warning("重载系统识别缓存失败", exc_info=True)


def _reload_knowledge_service_system_cache() -> None:
    try:
        from backend.service.knowledge_service import get_knowledge_service

        knowledge_service = get_knowledge_service()
        if hasattr(knowledge_service, "_system_list"):
            knowledge_service._system_list = None
    except Exception:
        logger.warning("重载知识库系统清单缓存失败", exc_info=True)


def _reload_related_caches() -> None:
    _reload_system_identification_cache()
    _reload_knowledge_service_system_cache()


def _upsert_systems(current_systems: List[Dict[str, Any]], incoming_systems: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    merged: List[Dict[str, Any]] = [dict(item) for item in current_systems]
    index_by_id: Dict[str, int] = {}
    index_by_name: Dict[str, int] = {}

    for idx, item in enumerate(merged):
        system_id = str(item.get("id") or "").strip()
        system_name = str(item.get("name") or "").strip()
        if system_id:
            index_by_id[system_id] = idx
        if system_name:
            index_by_name[system_name] = idx

    for item in incoming_systems:
        system_id = str(item.get("id") or "").strip()
        system_name = str(item.get("name") or "").strip()
        target_idx = None
        if system_id and system_id in index_by_id:
            target_idx = index_by_id[system_id]
        elif system_name and system_name in index_by_name:
            target_idx = index_by_name[system_name]

        if target_idx is None:
            merged.append(item)
            idx = len(merged) - 1
            if system_id:
                index_by_id[system_id] = idx
            if system_name:
                index_by_name[system_name] = idx
            continue

        merged[target_idx] = item
        if system_id:
            index_by_id[system_id] = target_idx
        if system_name:
            index_by_name[system_name] = target_idx

    return merged


def _extract_preview_errors(systems: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    valid_systems: List[Dict[str, Any]] = []
    preview_errors: List[Dict[str, Any]] = []

    for index, item in enumerate(systems or [], start=1):
        raw_errors = item.get("errors") if isinstance(item, dict) and isinstance(item.get("errors"), list) else []
        normalized_errors = [_cell_to_text(error) for error in raw_errors if _cell_to_text(error)]
        if normalized_errors:
            preview_errors.append(
                {
                    "row_number": _cell_to_int(item.get("row_number")) or index,
                    "system_name": _cell_to_text(item.get("name")),
                    "abbreviation": _cell_to_text(item.get("abbreviation")),
                    "errors": normalized_errors,
                }
            )
            continue
        valid_systems.append(item)

    return valid_systems, preview_errors


def _build_fallback_template() -> BytesIO:
    try:
        from openpyxl import Workbook
    except Exception as exc:
        raise ApiError(
            error_code="CATALOG_001",
            message="系统清单文件不合法",
            status_code=500,
            details={"reason": "缺少Excel解析依赖 openpyxl"},
        ) from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "应用系统清单"
    ws.append(["系统名称", "英文简称", "状态", "owner_id", "owner_username", "功能描述", "关联系统"])
    ws.append(["示例：统一支付平台", "PAY", "运行中", "user_demo_owner", "pm_demo", "支付统一受理", "核心账务,柜面渠道"])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@router.get("/template", dependencies=[Depends(require_system_list_admin)])
async def download_template(request: Request):
    if os.path.exists(TEMPLATE_PATH):
        try:
            from openpyxl import load_workbook

            with open(TEMPLATE_PATH, "rb") as template_file:
                wb = load_workbook(BytesIO(_sanitize_xlsx_for_openpyxl(template_file.read())), data_only=False)

            ws = wb.active
            header_values = [_normalize_header(cell) for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])]
            if "owner_id" not in header_values:
                ws.cell(row=1, column=len(header_values) + 1, value="owner_id")
                header_values.append("owner_id")
            if "owner_username" not in header_values:
                ws.cell(row=1, column=len(header_values) + 1, value="owner_username")

            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": 'attachment; filename="syslist-template.xlsx"'},
            )
        except Exception:
            logger.warning("读取模板文件失败，使用兜底模板", exc_info=True)

    try:
        output = _build_fallback_template()
    except ApiError as api_error:
        return build_error_response(
            request=request,
            status_code=api_error.status_code,
            error_code=api_error.error_code,
            message=api_error.message,
            details=api_error.details,
        )

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="syslist-template.xlsx"'},
    )


@router.post("/batch-import", dependencies=[Depends(require_system_list_admin)])
@router.post("/batch-import/preview", dependencies=[Depends(require_system_list_admin)])
async def batch_import_preview(request: Request, file: UploadFile = File(...)):
    try:
        content = await file.read()
        if not content:
            raise ApiError(
                error_code="CATALOG_001",
                message="系统清单文件不合法",
                status_code=400,
                details={"reason": "文件内容为空"},
            )

        wb = _load_workbook(content)
        systems = _parse_systems_sheet(_detect_system_sheet(wb))
        return {
            "code": 200,
            "data": {
                "systems": systems,
                "summary": {
                    "systems_total": len(systems),
                    "systems_error": sum(1 for item in systems if item.get("errors")),
                },
            },
        }
    except Exception as exc:
        api_error = _to_api_error(exc, default_error_code="CATALOG_001")
        return build_error_response(
            request=request,
            status_code=api_error.status_code,
            error_code=api_error.error_code,
            message=api_error.message,
            details=api_error.details,
        )


@router.post("/batch-import/confirm", dependencies=[Depends(require_system_list_admin)])
async def batch_import_confirm(request: Request, payload: SystemListImportConfirmRequest):
    mode = (payload.mode or "replace").strip().lower()
    if mode not in {"replace", "upsert"}:
        return build_error_response(
            request=request,
            status_code=400,
            error_code="CATALOG_001",
            message="导入模式不支持",
            details={"mode": mode, "supported_modes": ["replace", "upsert"]},
        )

    def _normalize_value(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    valid_payload_systems, preview_errors = _extract_preview_errors(payload.systems or [])
    systems: List[Dict[str, Any]] = []
    for item in valid_payload_systems:
        systems.append(
            {
                "id": item.get("id") or uuid.uuid4().hex,
                "name": _normalize_value(item.get("name")),
                "abbreviation": _normalize_value(item.get("abbreviation")),
                "status": _normalize_value(item.get("status")),
                "extra": _normalize_system_extra(item.get("extra")),
            }
        )

    try:
        write_applied = False
        if mode == "replace":
            if systems or not preview_errors:
                final_systems = systems
                write_applied = True
            else:
                final_systems = system_routes._read_systems()
        else:
            current_systems = system_routes._read_systems()
            final_systems = _upsert_systems(current_systems, systems)
            write_applied = bool(systems)

        if write_applied:
            system_routes._write_systems(final_systems)
            _reload_related_caches()

        execution_id = ""
        result_status = "partial_success" if preview_errors else "success"
        catalog_import_result = {
            "preview_errors": preview_errors,
            "updated_system_ids": [],
            "updated_systems": [],
            "skipped_items": [],
            "errors": [],
        }

        if (
            systems
            and getattr(settings, "ENABLE_V27_RUNTIME", False)
            and getattr(settings, "ENABLE_SYSTEM_CATALOG_PROFILE_INIT", False)
        ):
            init_result = get_system_catalog_profile_initializer().initialize_catalog(
                systems=systems,
                actor={"username": "admin"},
            )
            execution_id = str(init_result.get("execution_id") or "")
            if init_result.get("status") == "partial_success" or preview_errors:
                result_status = "partial_success"
            else:
                result_status = "success"
            catalog_import_result = {
                "preview_errors": preview_errors,
                "updated_system_ids": list(init_result.get("updated_system_ids") or []),
                "updated_systems": list(init_result.get("updated_systems") or []),
                "skipped_items": list(init_result.get("skipped_items") or []),
                "errors": list(init_result.get("errors") or []),
            }

        return {
            "code": 200,
            "message": "success",
            "result_status": result_status,
            "execution_id": execution_id,
            "catalog_import_result": catalog_import_result,
        }
    except Exception as exc:
        api_error = _to_api_error(exc, default_error_code="CATALOG_001")
        return build_error_response(
            request=request,
            status_code=api_error.status_code,
            error_code=api_error.error_code,
            message=api_error.message,
            details=api_error.details,
        )
