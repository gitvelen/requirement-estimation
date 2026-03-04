import json
from io import BytesIO

from openpyxl import Workbook

from backend.config.config import settings
from backend.service.esb_service import EsbService


class DummyEmbeddingService:
    def generate_embedding(self, text):
        return [1.0, 0.0, 0.0]

    def batch_generate_embeddings(self, texts, batch_size=25):
        return [[1.0, 0.0, 0.0] for _ in texts]


def build_service(tmp_path, monkeypatch):
    monkeypatch.setattr(settings, "REPORT_DIR", str(tmp_path))
    service = EsbService()
    service.embedding_service = DummyEmbeddingService()
    return service


def test_import_search_scope_and_status(tmp_path, monkeypatch):
    service = build_service(tmp_path, monkeypatch)

    csv_content = (
        "提供方系统简称,提供方中文名称,调用方系统简称,调用方中文名称,交易名称,交易码,服务场景码,状态\n"
        "SYS_A,系统A,SYS_B,系统B,白名单查询,SOFP500001,SC001,正常使用\n"
        "SYS_C,系统C,SYS_A,系统A,同步接口,SOFP500002,SC002,废弃使用\n"
    ).encode("utf-8")

    result = service.import_esb(csv_content, "esb.csv")
    assert result["imported"] == 2

    stats = service.get_stats(system_id="SYS_A")
    assert stats["active_entry_count"] == 1
    assert stats["deprecated_entry_count"] == 1

    provider_results = service.search_esb(
        query="接口",
        system_id="SYS_A",
        scope="provider",
        include_deprecated=False,
        similarity_threshold=0.0,
        top_k=5,
    )
    assert len(provider_results) == 1
    assert provider_results[0]["service_name"] == "白名单查询"

    consumer_results = service.search_esb(
        query="接口",
        system_id="SYS_A",
        scope="consumer",
        include_deprecated=True,
        similarity_threshold=0.0,
        top_k=5,
    )
    assert any(item["service_name"] == "同步接口" for item in consumer_results)


def test_import_xlsx_with_two_row_header_and_duplicate_system_id_columns(tmp_path, monkeypatch):
    """测试使用固定表头导入XLSX文件"""
    service = build_service(tmp_path, monkeypatch)

    wb = Workbook()
    ws = wb.active
    ws.title = "sheet1"

    # Row 1: 分组表头（会被跳过）
    ws["A1"] = "#"
    ws["B1"] = "投产日期"
    ws["O1"] = "调用方"

    # Row 2: 实际列名（会被跳过，使用固定表头）
    ws["C2"] = "系统标识"
    ws["D2"] = "系统名称"
    ws["H2"] = "服务名称"
    ws["I2"] = "场景名称"
    ws["J2"] = "交易码"
    ws["K2"] = "交易名称"
    ws["L2"] = "消费方系统标识"
    ws["M2"] = "消费方系统名称"

    # Row 3: 数据行（从第3行开始读取，按固定表头顺序）
    # 固定表头: 序号(A), 投产日期(B), 系统标识(C), 系统名称(D), 系统负责人(E),
    #          服务场景码(F), 服务名称(G), 场景名称(H), 交易码(I), 交易名称(J),
    #          消费方系统标识(K), 消费方系统名称(L), ...
    ws["A3"] = 1
    ws["B3"] = "2026-01-01"
    ws["C3"] = "SYS_A"  # 系统标识
    ws["D3"] = "系统A"  # 系统名称
    ws["J3"] = "账户查询"  # 交易名称（第10列）
    ws["K3"] = "SYS_B"  # 消费方系统标识（第11列）
    ws["L3"] = "系统B"  # 消费方系统名称（第12列）

    buf = BytesIO()
    wb.save(buf)

    result = service.import_esb(buf.getvalue(), "interface_template.xlsx")

    assert result["imported"] == 1
    assert result["mapping_resolved"]["provider_system_id"] == "系统标识"
    assert result["mapping_resolved"]["consumer_system_id"] == "消费方系统标识"
    assert result["mapping_resolved"]["service_name"] == "交易名称"

    with open(service.store_path, "r", encoding="utf-8") as f:
        store = json.load(f)

    entry = store["entries"][0]
    assert entry["provider_system_id"] == "SYS_A"
    assert entry["consumer_system_id"] == "SYS_B"
    assert entry["service_name"] == "账户查询"
    assert entry["status"] == "正常使用"
