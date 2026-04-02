import signal
from contextlib import contextmanager
from io import BytesIO

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

from backend.service.document_parser import DocumentParser


@contextmanager
def _timeout_guard(seconds: int):
    def _handle_timeout(_signum, _frame):
        raise TimeoutError(f"parse exceeded {seconds}s")

    previous = signal.signal(signal.SIGALRM, _handle_timeout)
    signal.setitimer(signal.ITIMER_REAL, seconds)
    try:
        yield
    finally:
        signal.setitimer(signal.ITIMER_REAL, 0)
        signal.signal(signal.SIGALRM, previous)


def test_parse_xlsx_accepts_workbook_with_datavalidation_id():
    parser = DocumentParser()

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["系统名称"])
    ws.append(["统一支付平台"])
    validation = DataValidation(type="list", formula1='"A,B"')
    ws.add_data_validation(validation)
    validation.add("A2")
    buf = BytesIO()
    wb.save(buf)

    parsed = parser.parse(
        file_content=buf.getvalue(),
        filename="syslist-template.xlsx",
        file_type="xlsx",
    )

    assert "Sheet1" in parsed
    assert len(parsed["Sheet1"]) > 1


def test_parse_xlsx_handles_sparse_sheet_without_timeout():
    parser = DocumentParser()

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["系统名称", "服务场景码", "交易名称", "消费方系统名称", "状态"])
    ws.append(["统一支付平台", "SC001", "支付查询", "核心账务", "正常使用"])
    ws.cell(row=200000, column=1, value="")

    buf = BytesIO()
    wb.save(buf)

    with _timeout_guard(3):
        parsed = parser.parse(
            file_content=buf.getvalue(),
            filename="接口治理台账.xlsx",
            file_type="xlsx",
        )

    assert list(parsed) == ["Sheet1"]
    assert len(parsed["Sheet1"]) == 2
