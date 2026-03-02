import os
import sys
from datetime import datetime
from pathlib import Path
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

ROOT_DIR = os.path.dirname(os.path.dirname(__file__))
if ROOT_DIR not in sys.path:
    sys.path.insert(0, ROOT_DIR)

from backend.app import app
from backend.api import routes as task_routes
from backend.config.config import settings
from backend.service import user_service


@pytest.fixture()
def client(tmp_path, monkeypatch):
    data_dir = tmp_path / "data"
    data_dir.mkdir(parents=True, exist_ok=True)

    monkeypatch.setattr(settings, "REPORT_DIR", str(data_dir))
    monkeypatch.setattr(settings, "DEBUG", False)

    monkeypatch.setattr(user_service, "USER_STORE_PATH", str(data_dir / "users.json"))
    monkeypatch.setattr(user_service, "USER_STORE_LOCK_PATH", str(data_dir / "users.json.lock"))

    monkeypatch.setattr(task_routes, "TASK_STORE_PATH", str(data_dir / "task_storage.json"))
    monkeypatch.setattr(task_routes, "TASK_STORE_LOCK_PATH", str(data_dir / "task_storage.json.lock"))

    return TestClient(app)


def _seed_user(username: str, password: str, roles):
    user = user_service.create_user_record(
        {
            "username": username,
            "display_name": username,
            "password": password,
            "roles": roles,
        }
    )
    with user_service.user_storage_context() as users:
        users.append(user)
    return user


def _login(client: TestClient, username: str, password: str) -> str:
    response = client.post("/api/v1/auth/login", json={"username": username, "password": password})
    assert response.status_code == 200
    return response.json()["data"]["token"]


def _seed_report_task(task_id: str, manager, expert, report_path: str):
    with task_routes._task_storage_context() as data:
        data[task_id] = {
            "task_id": task_id,
            "name": "报告下载任务",
            "creator_id": manager["id"],
            "creator_name": manager["display_name"],
            "status": "completed",
            "workflow_status": "completed",
            "created_at": datetime.now().isoformat(),
            "report_versions": [
                {"id": "rep_1", "round": 1, "version": 1, "file_path": report_path}
            ],
            "systems_data": {
                "HOP": [
                    {
                        "id": "feat_success",
                        "序号": "1.1",
                        "功能模块": "开户",
                        "功能点": "证件校验",
                        "业务描述": "校验身份证与OCR结果",
                        "输入": "身份证号",
                        "输出": "校验结果",
                        "依赖": "OCR服务",
                        "复杂度": "中",
                        "备注": "",
                        "预估人天": 2.58,
                        "optimistic": 1.5,
                        "most_likely": 2.5,
                        "pessimistic": 4.0,
                        "expected": 2.58,
                        "reasoning": "规则与接口复杂度中等",
                        "original_estimate": 2.5,
                    },
                    {
                        "id": "feat_degraded",
                        "序号": "1.2",
                        "功能模块": "销户",
                        "功能点": "销户申请",
                        "业务描述": "提交销户申请",
                        "输入": "申请单",
                        "输出": "申请结果",
                        "依赖": "核心系统",
                        "复杂度": "低",
                        "备注": "",
                        "预估人天": 1.5,
                        "optimistic": None,
                        "most_likely": None,
                        "pessimistic": None,
                        "expected": 1.5,
                        "reasoning": None,
                        "original_estimate": 1.5,
                    },
                ]
            },
            "expert_assignments": [
                {
                    "assignment_id": "assign_1",
                    "expert_id": expert["username"],
                    "expert_name": expert["display_name"],
                    "invite_token": "token_1",
                    "status": "submitted",
                    "round_submissions": {"1": datetime.now().isoformat()},
                }
            ],
        }


def test_report_download_expert_permissions_and_format(client, tmp_path):
    manager = _seed_user("rep_mgr", "pwd123", ["manager"])
    expert = _seed_user("rep_exp", "pwd123", ["expert"])
    outsider = _seed_user("rep_out", "pwd123", ["expert"])

    manager_token = _login(client, "rep_mgr", "pwd123")
    expert_token = _login(client, "rep_exp", "pwd123")
    outsider_token = _login(client, "rep_out", "pwd123")

    report_file = tmp_path / "report.pdf"
    report_file.write_bytes(b"%PDF-1.4\n%test")

    _seed_report_task("task_report", manager, expert, str(report_file))

    docx_resp = client.get(
        "/api/v1/tasks/task_report/report",
        params={"format": "docx"},
        headers={"Authorization": f"Bearer {manager_token}"},
    )
    assert docx_resp.status_code == 400
    assert docx_resp.json().get("error_code") == "REPORT_002"

    denied = client.get(
        "/api/v1/tasks/task_report/report",
        headers={"Authorization": f"Bearer {outsider_token}"},
    )
    assert denied.status_code == 403
    assert denied.json().get("error_code") == "AUTH_001"

    allowed = client.get(
        "/api/v1/tasks/task_report/report",
        headers={"Authorization": f"Bearer {expert_token}"},
    )
    assert allowed.status_code == 200
    assert "application/pdf" in str(allowed.headers.get("content-type") or "")


def test_report_download_missing_report_returns_report003(client, tmp_path):
    manager = _seed_user("rep_mgr2", "pwd123", ["manager"])
    token = _login(client, "rep_mgr2", "pwd123")

    with task_routes._task_storage_context() as data:
        data["task_no_report"] = {
            "task_id": "task_no_report",
            "name": "无报告任务",
            "creator_id": manager["id"],
            "creator_name": manager["display_name"],
            "status": "completed",
            "workflow_status": "evaluating",
            "created_at": datetime.now().isoformat(),
            "report_versions": [],
            "expert_assignments": [],
        }

    response = client.get(
        "/api/v1/tasks/task_no_report/report",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert response.status_code == 400
    assert response.json().get("error_code") == "REPORT_003"


def test_report_excel_export_contains_three_point_columns_and_degraded_values(client, tmp_path):
    manager = _seed_user("rep_mgr3", "pwd123", ["manager"])
    expert = _seed_user("rep_exp3", "pwd123", ["expert"])
    outsider = _seed_user("rep_out3", "pwd123", ["manager"])

    manager_token = _login(client, "rep_mgr3", "pwd123")
    outsider_token = _login(client, "rep_out3", "pwd123")

    report_file = tmp_path / "report_v24.pdf"
    report_file.write_bytes(b"%PDF-1.4\n%test")
    _seed_report_task("task_report_v24", manager, expert, str(report_file))

    denied = client.get(
        "/api/v1/reports/rep_1/export",
        headers={"Authorization": f"Bearer {outsider_token}"},
    )
    assert denied.status_code == 403
    assert denied.json().get("error_code") == "AUTH_001"

    resp = client.get(
        "/api/v1/reports/rep_1/export",
        headers={"Authorization": f"Bearer {manager_token}"},
    )
    assert resp.status_code == 200
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in str(resp.headers.get("content-type") or "")

    wb = load_workbook(filename=BytesIO(resp.content))
    ws = wb.active

    headers = [ws.cell(row=1, column=idx).value for idx in range(1, 17)]
    assert headers == [
        "系统",
        "序号",
        "功能模块",
        "功能点",
        "业务描述",
        "输入",
        "输出",
        "依赖项",
        "复杂度等级",
        "备注",
        "预估人天参考",
        "optimistic",
        "most_likely",
        "pessimistic",
        "expected",
        "reasoning",
    ]

    row_map = {
        ws.cell(row=row_idx, column=4).value: [ws.cell(row=row_idx, column=col).value for col in range(1, 17)]
        for row_idx in range(2, ws.max_row + 1)
    }

    success_row = row_map["证件校验"]
    assert success_row[11] == 1.5
    assert success_row[12] == 2.5
    assert success_row[13] == 4.0
    assert success_row[14] == 2.58
    assert success_row[15] == "规则与接口复杂度中等"

    degraded_row = row_map["销户申请"]
    assert degraded_row[11] == "N/A"
    assert degraded_row[12] == "N/A"
    assert degraded_row[13] == "N/A"
    assert degraded_row[14] == "N/A"
    assert degraded_row[15] == "LLM 估算未成功"
