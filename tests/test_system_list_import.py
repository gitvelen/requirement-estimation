import os
import sys
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

ROOT_DIR = os.path.dirname(os.path.dirname(__file__))
if ROOT_DIR not in sys.path:
    sys.path.insert(0, ROOT_DIR)

from backend.app import app
from backend.api import system_routes
from backend.config.config import settings
from backend.service import memory_service
from backend.service import runtime_execution_service
from backend.service import system_profile_service
from backend.service import user_service


@pytest.fixture()
def client(tmp_path, monkeypatch):
    data_dir = tmp_path / "data"
    data_dir.mkdir(parents=True, exist_ok=True)

    monkeypatch.setattr(settings, "REPORT_DIR", str(data_dir))
    monkeypatch.setattr(user_service, "USER_STORE_PATH", str(data_dir / "users.json"))
    monkeypatch.setattr(user_service, "USER_STORE_LOCK_PATH", str(data_dir / "users.json.lock"))
    monkeypatch.setattr(system_routes, "CSV_PATH", str(tmp_path / "system_list.csv"))

    system_profile_service._system_profile_service = None
    memory_service._memory_service = None
    runtime_execution_service._runtime_execution_service = None

    return TestClient(app)


def _seed_admin():
    admin = user_service.create_user_record(
        {
            "username": "admin",
            "display_name": "管理员",
            "password": "admin123",
            "roles": ["admin"],
        }
    )
    with user_service.user_storage_context() as users:
        users.append(admin)
    return admin


def _seed_manager(username: str = "manager", password: str = "manager123", display_name: str = "经理"):
    manager = user_service.create_user_record(
        {
            "username": username,
            "display_name": display_name,
            "password": password,
            "roles": ["manager"],
        }
    )
    with user_service.user_storage_context() as users:
        users.append(manager)
    return manager


def _login(client: TestClient, username: str, password: str) -> str:
    response = client.post("/api/v1/auth/login", json={"username": username, "password": password})
    assert response.status_code == 200
    return response.json()["data"]["token"]


def _build_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "应用系统清单"
    ws.append(["系统名称", "英文简称", "状态", "功能描述", "关联系统"])
    ws.append(["HOP", "HOP", "运行中", "核心系统", "CLMP,柜面"])
    ws.append(["CLMP", "CLMP", "建设中", "贷款平台", ""])

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _build_template_with_invalid_row_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "应用系统清单"
    ws.append(["系统名称", "英文简称", "状态", "功能描述", "关联系统"])
    ws.append(["HOP", "HOP", "运行中", "核心系统", "CLMP,柜面"])
    ws.append(["老营销系统", "", "已下线", "", ""])

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _build_owner_alias_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "应用系统清单"
    ws.append(["系统名称", "英文简称", "状态", "系统负责人ID", "系统负责人账号", "负责人姓名"])
    ws.append(["HOP", "HOP", "运行中", "", "manager", "张三"])

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _build_legacy_owner_name_xlsx(owner_name: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "应用系统清单"
    ws.append(["系统名称", "英文简称", "状态", "系统负责人"])
    ws.append(["NESB", "NESB", "运行中", owner_name])

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def test_system_list_batch_import_replace(client, monkeypatch):
    monkeypatch.setattr(settings, "DEBUG", False)
    monkeypatch.setattr(settings, "ENABLE_V27_PROFILE_SCHEMA", True)
    monkeypatch.setattr(settings, "ENABLE_V27_RUNTIME", True)
    monkeypatch.setattr(settings, "ENABLE_SYSTEM_CATALOG_PROFILE_INIT", True)

    _seed_admin()
    token = _login(client, "admin", "admin123")

    excel_bytes = _build_template_xlsx()
    preview = client.post(
        "/api/v1/system-list/batch-import/preview",
        files={"file": ("system_list.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": f"Bearer {token}"},
    )
    assert preview.status_code == 200
    data = preview.json()["data"]
    assert data["summary"] == {
        "systems_total": 2,
        "systems_error": 0,
    }
    assert len(data["systems"]) == 2

    confirm = client.post(
        "/api/v1/system-list/batch-import/confirm",
        json={"mode": "replace", "systems": data["systems"]},
        headers={"Authorization": f"Bearer {token}"},
    )
    assert confirm.status_code == 200

    payload = confirm.json()
    assert payload["message"] == "success"
    assert payload["result_status"] == "success"
    assert payload["execution_id"]
    assert payload["catalog_import_result"]["skipped_items"] == []
    assert len(payload["catalog_import_result"]["updated_system_ids"]) == 2

    systems = system_routes._read_systems()
    assert {item["name"] for item in systems} == {"HOP", "CLMP"}
    assert {item["abbreviation"] for item in systems} == {"HOP", "CLMP"}
    assert all(item.get("id") for item in systems)
    by_name = {item["name"]: item for item in systems}
    assert by_name["HOP"].get("extra", {}).get("功能描述") == "核心系统"
    assert by_name["HOP"].get("extra", {}).get("关联系统") == "CLMP,柜面"
    assert by_name["CLMP"].get("extra", {}).get("功能描述") == "贷款平台"


def test_system_list_preview_owner_alias_and_auth_error_schema(client, monkeypatch):
    monkeypatch.setattr(settings, "DEBUG", False)

    _seed_admin()
    _seed_manager()
    token = _login(client, "admin", "admin123")

    excel_bytes = _build_owner_alias_xlsx()
    preview = client.post(
        "/api/v1/system-list/batch-import/preview",
        files={"file": ("system_list_alias.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": f"Bearer {token}", "X-Request-ID": "req_system_list_alias"},
    )
    assert preview.status_code == 200
    data = preview.json()["data"]
    assert data["summary"]["systems_total"] == 1
    extra = data["systems"][0].get("extra", {})
    assert extra.get("owner_id") == ""
    assert extra.get("owner_username") == "manager"
    assert extra.get("owner_name") == "张三"
    assert "系统负责人ID" not in extra

    confirm = client.post(
        "/api/v1/system-list/batch-import/confirm",
        json={"mode": "replace", "systems": data["systems"]},
        headers={"Authorization": f"Bearer {token}"},
    )
    assert confirm.status_code == 200

    owner_info = system_routes.resolve_system_owner(system_name="HOP")
    assert owner_info.get("system_found") is True
    assert owner_info.get("resolved_by") == "owner_username"
    assert owner_info.get("resolved_owner_id")

    unauthorized = client.post(
        "/api/v1/system-list/batch-import/preview",
        files={"file": ("system_list_alias.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": "Bearer invalid", "X-Request-ID": "req_system_list_auth"},
    )
    assert unauthorized.status_code == 403
    payload = unauthorized.json()
    assert payload.get("error_code") == "AUTH_001"
    assert payload.get("message")
    assert payload.get("request_id") == "req_system_list_auth"


def test_system_list_confirm_returns_preview_errors_and_skips_invalid_rows(client, monkeypatch):
    monkeypatch.setattr(settings, "DEBUG", False)
    monkeypatch.setattr(settings, "ENABLE_V27_PROFILE_SCHEMA", True)
    monkeypatch.setattr(settings, "ENABLE_V27_RUNTIME", True)
    monkeypatch.setattr(settings, "ENABLE_SYSTEM_CATALOG_PROFILE_INIT", True)

    _seed_admin()
    token = _login(client, "admin", "admin123")

    excel_bytes = _build_template_with_invalid_row_xlsx()
    preview = client.post(
        "/api/v1/system-list/batch-import/preview",
        files={"file": ("system_list_invalid.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": f"Bearer {token}"},
    )
    assert preview.status_code == 200

    data = preview.json()["data"]
    assert data["summary"] == {
        "systems_total": 2,
        "systems_error": 1,
    }
    assert data["systems"][1]["row_number"] == 3
    assert data["systems"][1]["errors"] == ["英文简称不能为空"]

    confirm = client.post(
        "/api/v1/system-list/batch-import/confirm",
        json={"mode": "replace", "systems": data["systems"]},
        headers={"Authorization": f"Bearer {token}"},
    )
    assert confirm.status_code == 200

    payload = confirm.json()
    assert payload["result_status"] == "partial_success"
    assert payload["catalog_import_result"]["preview_errors"] == [
        {
            "row_number": 3,
            "system_name": "老营销系统",
            "abbreviation": "",
            "errors": ["英文简称不能为空"],
        }
    ]
    assert len(payload["catalog_import_result"]["updated_system_ids"]) == 1
    assert payload["catalog_import_result"]["skipped_items"] == []
    assert payload["catalog_import_result"]["errors"] == []

    systems = system_routes._read_systems()
    assert {item["name"] for item in systems} == {"HOP"}
    assert {item["abbreviation"] for item in systems} == {"HOP"}


def test_system_list_legacy_owner_name_can_resolve_manager_display_name(client, monkeypatch):
    monkeypatch.setattr(settings, "DEBUG", False)

    _seed_admin()
    manager = _seed_manager(display_name="黄洋")
    token = _login(client, "admin", "admin123")

    excel_bytes = _build_legacy_owner_name_xlsx("黄洋")
    preview = client.post(
        "/api/v1/system-list/batch-import/preview",
        files={"file": ("system_list_legacy_owner.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": f"Bearer {token}", "X-Request-ID": "req_system_list_legacy_owner"},
    )
    assert preview.status_code == 200
    data = preview.json()["data"]
    assert data["summary"]["systems_total"] == 1
    extra = data["systems"][0].get("extra", {})
    assert extra.get("owner_name") == "黄洋"

    confirm = client.post(
        "/api/v1/system-list/batch-import/confirm",
        json={"mode": "replace", "systems": data["systems"]},
        headers={"Authorization": f"Bearer {token}"},
    )
    assert confirm.status_code == 200

    owner_info = system_routes.resolve_system_owner(system_name="NESB")
    assert owner_info.get("system_found") is True
    assert owner_info.get("resolved_owner_id") == manager["id"]


def test_system_list_confirm_invalid_mode_returns_catalog_error(client, monkeypatch):
    monkeypatch.setattr(settings, "DEBUG", False)

    _seed_admin()
    token = _login(client, "admin", "admin123")

    response = client.post(
        "/api/v1/system-list/batch-import/confirm",
        json={"mode": "invalid_mode", "systems": []},
        headers={"Authorization": f"Bearer {token}", "X-Request-ID": "req_system_list_mode"},
    )

    assert response.status_code == 400
    payload = response.json()
    assert payload.get("error_code") == "CATALOG_001"
    assert payload.get("request_id") == "req_system_list_mode"
