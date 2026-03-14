import os
import sys
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

ROOT_DIR = os.path.dirname(os.path.dirname(__file__))
if ROOT_DIR not in sys.path:
    sys.path.insert(0, ROOT_DIR)

from backend.app import app
from backend.api import system_routes
from backend.config.config import settings
from backend.service import esb_service
from backend.service import memory_service
from backend.service import runtime_execution_service
from backend.service import system_profile_service
from backend.service import user_service


class DummyEmbeddingService:
    def generate_embedding(self, text):
        return [1.0, 0.0, 0.0]

    def batch_generate_embeddings(self, texts, batch_size=25):
        return [[1.0, 0.0, 0.0] for _ in texts]


class BrokenEmbeddingService:
    def generate_embedding(self, text):
        raise RuntimeError("embedding unavailable")

    def batch_generate_embeddings(self, texts, batch_size=25):
        raise RuntimeError("embedding unavailable")


@pytest.fixture()
def client(tmp_path, monkeypatch):
    data_dir = tmp_path / "data"
    data_dir.mkdir(parents=True, exist_ok=True)

    monkeypatch.setattr(settings, "REPORT_DIR", str(data_dir))
    monkeypatch.setattr(settings, "DEBUG", False)

    monkeypatch.setattr(user_service, "USER_STORE_PATH", str(data_dir / "users.json"))
    monkeypatch.setattr(user_service, "USER_STORE_LOCK_PATH", str(data_dir / "users.json.lock"))

    monkeypatch.setattr(system_routes, "CSV_PATH", str(tmp_path / "system_list.csv"))

    esb_service._esb_service = None
    system_profile_service._system_profile_service = None
    memory_service._memory_service = None
    runtime_execution_service._runtime_execution_service = None

    return TestClient(app)


def _seed_user(username: str, password: str, roles):
    user = user_service.create_user_record(
        {
            "username": username,
            "display_name": username,
            "password": password,
            "roles": roles,
        }
    )
    with user_service.user_storage_context() as users:
        users.append(user)
    return user


def _login(client: TestClient, username: str, password: str) -> str:
    response = client.post("/api/v1/auth/login", json={"username": username, "password": password})
    assert response.status_code == 200
    return response.json()["data"]["token"]


def _seed_system_owner(system_name: str, system_id: str, owner_id: str, abbreviation: str | None = None):
    system_routes._write_systems(
        [
            {
                "id": system_id,
                "name": system_name,
                "abbreviation": abbreviation or system_name,
                "status": "运行中",
                "extra": {"owner_id": owner_id},
            }
        ]
    )


def _build_esb_workbook_with_offset_header() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "接口申请-20250724常规版本"

    for i in range(1, 6):
        ws.append([f"说明{i}", "", "", "", "", ""])

    ws.append(["系统标识", "系统名称", "服务场景码", "交易名称", "系统标识", "系统名称"])
    ws.append(["sys_hop", "HOP", "scene_01", "查询接口", "sys_x", "X系统"])

    extra = wb.create_sheet("字典")
    extra.append(["字段", "说明"])
    extra.append(["系统标识", "系统唯一ID"])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _build_esb_workbook_with_valid_and_unrelated_sheets() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "接口申请"
    ws.append(["系统标识", "系统名称", "服务场景码", "交易名称", "系统标识", "系统名称"])
    ws.append(["sys_other", "OTHER", "scene_01", "查询接口", "sys_x", "X系统"])

    unrelated = wb.create_sheet("新服务治理平台服务视图")
    unrelated.append(["序号", "模块", "说明"])
    unrelated.append(["1", "A", "B"])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _build_esb_workbook_with_deep_header_row() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "接口申请-深层表头"

    for i in range(1, 121):
        ws.append([f"说明{i}", "", "", "", "", ""])

    ws.append(["系统标识", "系统名称", "服务场景码", "交易名称", "系统标识", "系统名称"])
    ws.append(["sys_hop", "HOP", "scene_01", "查询接口", "sys_x", "X系统"])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_esb_import_requires_owner(client):
    owner = _seed_user("esb_owner", "owner123", ["manager"])
    other = _seed_user("esb_other", "other123", ["manager"])

    _seed_system_owner("HOP", "sys_hop", owner["id"])

    other_token = _login(client, "esb_other", "other123")

    csv_content = (
        "提供方系统简称,调用方系统简称,交易名称,状态\n"
        "sys_hop,sys_x,查询接口,正常使用\n"
    ).encode("utf-8")

    response = client.post(
        "/api/v1/esb/imports",
        data={"system_id": "sys_hop"},
        files={"file": ("esb.csv", csv_content, "text/csv")},
        headers={"Authorization": f"Bearer {other_token}", "X-Request-ID": "req_esb_auth"},
    )

    assert response.status_code == 403
    payload = response.json()
    assert payload["error_code"] == "AUTH_001"
    assert payload["request_id"] == "req_esb_auth"


def test_esb_import_missing_required_columns_returns_esb002(client):
    owner = _seed_user("esb_owner2", "owner123", ["manager"])
    _seed_system_owner("HOP", "sys_hop", owner["id"])
    token = _login(client, "esb_owner2", "owner123")

    invalid_csv = (
        "系统,接口名\n"
        "sys_hop,查询\n"
    ).encode("utf-8")

    response = client.post(
        "/api/v1/esb/imports",
        data={"system_id": "sys_hop"},
        files={"file": ("esb.csv", invalid_csv, "text/csv")},
        headers={"Authorization": f"Bearer {token}"},
    )

    assert response.status_code == 400
    assert response.json()["error_code"] == "ESB_002"


def test_esb_import_filters_by_system_id_and_updates_completeness(client):
    owner = _seed_user("esb_owner3", "owner123", ["manager"])
    _seed_system_owner("HOP", "sys_hop", owner["id"])
    token = _login(client, "esb_owner3", "owner123")

    service = esb_service.get_esb_service()
    service.embedding_service = DummyEmbeddingService()

    csv_content = (
        "提供方系统简称,调用方系统简称,交易名称,状态\n"
        "sys_hop,sys_x,查询接口,正常使用\n"
        "sys_other,sys_else,同步接口,正常使用\n"
    ).encode("utf-8")

    mapping_json = {
        "provider_system_id": ["提供方系统简称"],
        "consumer_system_id": ["调用方系统简称"],
        "service_name": "交易名称",
        "status": "状态",
    }

    response = client.post(
        "/api/v1/esb/imports",
        data={"system_id": "sys_hop", "mapping_json": str(mapping_json).replace("'", '"')},
        files={"file": ("esb.csv", csv_content, "text/csv")},
        headers={"Authorization": f"Bearer {token}"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["imported"] == 1
    assert payload["skipped"] >= 1
    assert payload.get("mapping_resolved", {}).get("provider_system_id") == "提供方系统简称"
    assert payload.get("mapping_resolved", {}).get("service_name") == "交易名称"

    profile = system_profile_service.get_system_profile_service().get_profile("HOP")
    assert profile
    assert profile.get("completeness", {}).get("esb") is True
    assert int(profile.get("completeness_score") or 0) >= 30


def test_esb_import_accepts_system_abbreviation_or_name_when_tab_uses_uuid_system_id(client):
    owner = _seed_user("esb_owner_uuid_alias", "owner123", ["manager"])
    system_id = "6d8a1fc0d67e4b7785f1a9a2670d08c6"
    _seed_system_owner("贷款核算", system_id, owner["id"], abbreviation="ULCA")
    token = _login(client, "esb_owner_uuid_alias", "owner123")

    service = esb_service.get_esb_service()
    service.embedding_service = DummyEmbeddingService()

    csv_content = (
        "提供方系统简称,调用方系统简称,交易名称,状态\n"
        "ULCA,sys_x,接口A,正常使用\n"
        "贷款核算,sys_y,接口B,正常使用\n"
        "OTHER,sys_z,接口C,正常使用\n"
    ).encode("utf-8")

    response = client.post(
        "/api/v1/esb/imports",
        data={
            "system_id": system_id,
            "mapping_provider_system_id": "提供方系统简称",
            "mapping_consumer_system_id": "调用方系统简称",
            "mapping_service_name": "交易名称",
            "mapping_status": "状态",
        },
        files={"file": ("esb_uuid_alias.csv", csv_content, "text/csv")},
        headers={"Authorization": f"Bearer {token}"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["imported"] == 2
    assert payload["skipped"] == 1

    profile = system_profile_service.get_system_profile_service().get_profile("贷款核算")
    assert profile
    assert profile.get("completeness", {}).get("esb") is True


def test_v27_esb_import_requires_admin_for_global_governance_mode(client, monkeypatch):
    monkeypatch.setattr(settings, "ENABLE_V27_PROFILE_SCHEMA", True)
    monkeypatch.setattr(settings, "ENABLE_V27_RUNTIME", True)
    monkeypatch.setattr(settings, "ENABLE_SERVICE_GOVERNANCE_IMPORT", True)

    manager = _seed_user("v27_governance_manager", "pwd123", ["manager"])
    token = _login(client, "v27_governance_manager", "pwd123")

    csv_content = (
        "系统名称,服务场景码,交易名称,消费方系统名称,状态\n"
        "统一支付平台,SC001,支付查询,核心账务,正常使用\n"
    ).encode("utf-8")

    response = client.post(
        "/api/v1/esb/imports",
        files={"file": ("governance.csv", csv_content, "text/csv")},
        headers={"Authorization": f"Bearer {token}"},
    )

    assert response.status_code == 403
    assert response.json()["error_code"] == "AUTH_001"


def test_v27_esb_import_returns_explicit_flag_error_when_global_governance_disabled(client, monkeypatch):
    monkeypatch.setattr(settings, "ENABLE_V27_PROFILE_SCHEMA", False)
    monkeypatch.setattr(settings, "ENABLE_V27_RUNTIME", False)
    monkeypatch.setattr(settings, "ENABLE_SERVICE_GOVERNANCE_IMPORT", False)

    _seed_user("v27_governance_admin", "pwd123", ["admin"])
    token = _login(client, "v27_governance_admin", "pwd123")

    csv_content = (
        "服务方系统名称,服务场景码,交易名称,消费方系统名称,状态\n"
        "办公自动化,2022000601,cus提交流程,智能报销,正常使用\n"
    ).encode("utf-8")

    response = client.post(
        "/api/v1/esb/imports",
        files={"file": ("governance.csv", csv_content, "text/csv")},
        headers={"Authorization": f"Bearer {token}"},
    )

    assert response.status_code == 400
    payload = response.json()
    assert payload["error_code"] == "ESB_002"
    assert payload["message"] == "服务治理全局导入未启用"
    assert payload["details"]["missing_flags"] == [
        "ENABLE_V27_PROFILE_SCHEMA",
        "ENABLE_V27_RUNTIME",
        "ENABLE_SERVICE_GOVERNANCE_IMPORT",
    ]


def test_esb_import_accepts_human_friendly_mapping_fields(client):
    owner = _seed_user("esb_owner_human_map", "owner123", ["manager"])
    _seed_system_owner("HOP", "sys_hop", owner["id"])
    token = _login(client, "esb_owner_human_map", "owner123")

    service = esb_service.get_esb_service()
    service.embedding_service = DummyEmbeddingService()

    csv_content = (
        "提供方系统简称,调用方系统简称,交易名称,状态\n"
        "sys_hop,sys_x,查询接口,正常使用\n"
    ).encode("utf-8")

    response = client.post(
        "/api/v1/esb/imports",
        data={
            "system_id": "sys_hop",
            "mapping_provider_system_id": "提供方系统简称",
            "mapping_consumer_system_id": "调用方系统简称",
            "mapping_service_name": "交易名称",
            "mapping_status": "状态",
        },
        files={"file": ("esb.csv", csv_content, "text/csv")},
        headers={"Authorization": f"Bearer {token}"},
    )
    assert response.status_code == 200
    payload = response.json()
    assert payload["imported"] == 1
    assert payload.get("mapping_resolved", {}).get("provider_system_id") == "提供方系统简称"
    assert payload.get("mapping_resolved", {}).get("consumer_system_id") == "调用方系统简称"
    assert payload.get("mapping_resolved", {}).get("service_name") == "交易名称"
    assert payload.get("mapping_resolved", {}).get("status") == "状态"


def test_esb_import_accepts_interface_template_header_beyond_first_five_rows(client):
    owner = _seed_user("esb_owner_offset_header", "owner123", ["manager"])
    _seed_system_owner("HOP", "sys_hop", owner["id"])
    token = _login(client, "esb_owner_offset_header", "owner123")

    service = esb_service.get_esb_service()
    service.embedding_service = DummyEmbeddingService()

    xlsx_content = _build_esb_workbook_with_offset_header()

    response = client.post(
        "/api/v1/esb/imports",
        data={"system_id": "sys_hop"},
        files={"file": ("esb_offset_header.xlsx", xlsx_content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": f"Bearer {token}"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["imported"] == 1
    assert payload.get("mapping_resolved", {}).get("provider_system_id") == "系统标识"
    assert payload.get("mapping_resolved", {}).get("consumer_system_id") == "系统标识#2"
    assert payload.get("mapping_resolved", {}).get("service_name") == "交易名称"


def test_esb_import_accepts_service_provider_consumer_aliases_without_status_column(client):
    owner = _seed_user("esb_owner_aliases", "owner123", ["manager"])
    _seed_system_owner("HOP", "sys_hop", owner["id"])
    token = _login(client, "esb_owner_aliases", "owner123")

    service = esb_service.get_esb_service()
    service.embedding_service = DummyEmbeddingService()

    csv_content = (
        "服务方系统标识,消费方系统标识,服务名称\n"
        "sys_hop,sys_x,查询接口\n"
    ).encode("utf-8")

    response = client.post(
        "/api/v1/esb/imports",
        data={"system_id": "sys_hop"},
        files={"file": ("esb_aliases.csv", csv_content, "text/csv")},
        headers={"Authorization": f"Bearer {token}"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["imported"] == 1
    assert payload.get("mapping_resolved", {}).get("provider_system_id") == "服务方系统标识"
    assert payload.get("mapping_resolved", {}).get("consumer_system_id") == "消费方系统标识"
    assert payload.get("mapping_resolved", {}).get("service_name") == "服务名称"


def test_esb_import_does_not_raise_missing_required_when_detail_sheet_exists_but_rows_filtered(client):
    owner = _seed_user("esb_owner_filtered_rows", "owner123", ["manager"])
    _seed_system_owner("HOP", "sys_hop", owner["id"])
    token = _login(client, "esb_owner_filtered_rows", "owner123")

    service = esb_service.get_esb_service()
    service.embedding_service = DummyEmbeddingService()

    xlsx_content = _build_esb_workbook_with_valid_and_unrelated_sheets()

    response = client.post(
        "/api/v1/esb/imports",
        data={"system_id": "sys_hop"},
        files={"file": ("esb_valid_plus_unrelated.xlsx", xlsx_content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": f"Bearer {token}"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["imported"] == 0
    assert payload["skipped"] >= 1
    assert payload.get("mapping_resolved", {}).get("provider_system_id") == "系统标识"


def test_esb_import_accepts_header_beyond_fifty_rows(client):
    owner = _seed_user("esb_owner_deep_header", "owner123", ["manager"])
    _seed_system_owner("HOP", "sys_hop", owner["id"])
    token = _login(client, "esb_owner_deep_header", "owner123")

    service = esb_service.get_esb_service()
    service.embedding_service = DummyEmbeddingService()

    xlsx_content = _build_esb_workbook_with_deep_header_row()

    response = client.post(
        "/api/v1/esb/imports",
        data={"system_id": "sys_hop"},
        files={"file": ("esb_deep_header.xlsx", xlsx_content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": f"Bearer {token}"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["imported"] == 1
    assert payload.get("mapping_resolved", {}).get("provider_system_id") == "系统标识"
    assert payload.get("mapping_resolved", {}).get("consumer_system_id") == "系统标识#2"


def test_esb_import_embedding_unavailable_returns_emb001(client):
    owner = _seed_user("esb_owner4", "owner123", ["manager"])
    _seed_system_owner("HOP", "sys_hop", owner["id"])
    token = _login(client, "esb_owner4", "owner123")

    service = esb_service.get_esb_service()
    service.embedding_service = BrokenEmbeddingService()

    csv_content = (
        "提供方系统简称,调用方系统简称,交易名称,状态\n"
        "sys_hop,sys_x,查询接口,正常使用\n"
    ).encode("utf-8")

    response = client.post(
        "/api/v1/esb/imports",
        data={"system_id": "sys_hop"},
        files={"file": ("esb.csv", csv_content, "text/csv")},
        headers={"Authorization": f"Bearer {token}"},
    )

    assert response.status_code == 503
    assert response.json()["error_code"] == "EMB_001"

    profile = system_profile_service.get_system_profile_service().get_profile("HOP")
    assert not profile or not profile.get("completeness", {}).get("esb")


def test_esb_search_and_stats_support_scope_and_include_deprecated(client):
    owner = _seed_user("esb_owner5", "owner123", ["manager"])
    _seed_system_owner("HOP", "sys_hop", owner["id"])
    token = _login(client, "esb_owner5", "owner123")

    service = esb_service.get_esb_service()
    service.embedding_service = DummyEmbeddingService()

    csv_content = (
        "提供方系统简称,调用方系统简称,交易名称,状态\n"
        "sys_hop,sys_x,查询接口,正常使用\n"
        "sys_y,sys_hop,同步接口,废弃使用\n"
    ).encode("utf-8")

    import_response = client.post(
        "/api/v1/esb/imports",
        data={"system_id": "sys_hop"},
        files={"file": ("esb.csv", csv_content, "text/csv")},
        headers={"Authorization": f"Bearer {token}"},
    )
    assert import_response.status_code == 200

    stats_response = client.get(
        "/api/v1/esb/stats",
        params={"system_id": "sys_hop"},
        headers={"Authorization": f"Bearer {token}"},
    )
    assert stats_response.status_code == 200
    stats_payload = stats_response.json()
    assert stats_payload["active_entry_count"] == 1
    assert stats_payload["deprecated_entry_count"] == 1
    assert stats_payload["active_unique_service_count"] == 1

    search_response = client.get(
        "/api/v1/esb/search",
        params={
            "q": "接口",
            "system_id": "sys_hop",
            "scope": "consumer",
            "include_deprecated": "false",
            "top_k": 10,
        },
        headers={"Authorization": f"Bearer {token}"},
    )
    assert search_response.status_code == 200
    search_payload = search_response.json()
    assert search_payload["total"] == 0
    assert search_payload["items"] == []

    search_with_deprecated_response = client.get(
        "/api/v1/esb/search",
        params={
            "q": "接口",
            "system_id": "sys_hop",
            "scope": "consumer",
            "include_deprecated": "true",
            "top_k": 10,
        },
        headers={"Authorization": f"Bearer {token}"},
    )
    assert search_with_deprecated_response.status_code == 200
    search_with_deprecated_payload = search_with_deprecated_response.json()
    assert search_with_deprecated_payload["total"] >= 1
    assert any(item.get("status") == "废弃使用" for item in search_with_deprecated_payload["items"])


def test_esb_search_requires_owner_or_admin(client):
    owner = _seed_user("esb_owner6", "owner123", ["manager"])
    other = _seed_user("esb_other6", "other123", ["manager"])
    _seed_system_owner("HOP", "sys_hop", owner["id"])
    token = _login(client, "esb_other6", "other123")

    response = client.get(
        "/api/v1/esb/search",
        params={"q": "接口", "system_id": "sys_hop"},
        headers={"Authorization": f"Bearer {token}"},
    )
    assert response.status_code == 403
    assert response.json()["error_code"] == "AUTH_001"
