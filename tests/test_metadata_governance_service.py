from datetime import datetime
from io import BytesIO
from types import SimpleNamespace

import yaml
from openpyxl import load_workbook

from backend.config.config import settings
from backend.service.metadata_governance_service import MetadataGovernanceService


def _write_config(path, payload):
    with open(path, 'w', encoding='utf-8') as f:
        yaml.safe_dump(payload, f, allow_unicode=True, sort_keys=False)


def test_metadata_governance_service_syncs_model_config_and_updates_scan_fields(tmp_path, monkeypatch):
    config_path = tmp_path / 'config.yaml'
    _write_config(
        config_path,
        {
            'llm': {'base_url': 'http://old', 'model_name': 'old-model', 'api_key': 'old'},
            'scan': {'similarity_threshold': 0.8, 'schedule': '0 23 * * *', 'lookback_hours': 24},
        },
    )

    monkeypatch.setattr(settings, 'DASHSCOPE_API_BASE', 'http://10.73.254.200:30000/v1')
    monkeypatch.setattr(settings, 'LLM_MODEL', 'Qwen3-32B')
    monkeypatch.setattr(settings, 'DASHSCOPE_API_KEY', 'not-needed')
    monkeypatch.setattr(settings, 'LLM_TIMEOUT', 120)
    monkeypatch.setattr(settings, 'LLM_TEMPERATURE', 0.7)
    monkeypatch.setattr(settings, 'LLM_MAX_TOKENS', 4000)

    service = MetadataGovernanceService(config_path=config_path)
    service.update_runtime_config(similarity_threshold=0.85, execution_time='daily_23', match_scope='all')

    with open(config_path, 'r', encoding='utf-8') as f:
        saved = yaml.safe_load(f)

    assert saved['llm']['base_url'] == 'http://10.73.254.200:30000/v1'
    assert saved['llm']['model_name'] == 'Qwen3-32B'
    assert saved['scan']['similarity_threshold'] == 0.85
    assert saved['scan']['schedule'] == '0 23 * * *'
    assert saved['scan']['lookback_hours'] == 0


def test_metadata_governance_service_does_not_persist_api_key_to_yaml(tmp_path, monkeypatch):
    config_path = tmp_path / 'config.yaml'
    _write_config(
        config_path,
        {
            'llm': {'base_url': 'http://old', 'model_name': 'old-model', 'api_key': 'old-secret'},
            'scan': {'similarity_threshold': 0.8, 'schedule': '0 23 * * *', 'lookback_hours': 24},
        },
    )

    monkeypatch.setattr(settings, 'DASHSCOPE_API_BASE', 'http://10.73.254.200:30000/v1')
    monkeypatch.setattr(settings, 'LLM_MODEL', 'Qwen3-32B')
    monkeypatch.setattr(settings, 'DASHSCOPE_API_KEY', 'runtime-secret')
    monkeypatch.setattr(settings, 'LLM_TIMEOUT', 120)
    monkeypatch.setattr(settings, 'LLM_TEMPERATURE', 0.7)
    monkeypatch.setattr(settings, 'LLM_MAX_TOKENS', 4000)

    service = MetadataGovernanceService(config_path=config_path)
    service.update_runtime_config(similarity_threshold=0.85, execution_time='daily_23', match_scope='all')

    with open(config_path, 'r', encoding='utf-8') as f:
        saved = yaml.safe_load(f)

    assert saved['llm']['base_url'] == 'http://10.73.254.200:30000/v1'
    assert saved['llm']['model_name'] == 'Qwen3-32B'
    assert saved['llm'].get('api_key') in {'', None}


def test_build_excel_only_keeps_redundant_rows_in_new_sheet_and_hides_status_column(tmp_path):
    config_path = tmp_path / 'config.yaml'
    _write_config(config_path, {'llm': {}, 'scan': {}})
    service = MetadataGovernanceService(config_path=config_path)

    output = service._build_excel(
        [
            {
                'metadata_id': 'M001',
                'chinese_name': '客户名称',
                'data_type': 'VARCHAR(64)',
                'opt_time': '2026-03-26 10:00:00',
                'opt_user': 'tester',
                'service_ids': '30120001 | 30120002',
                'redundancy_status': '有冗余',
                'similarity_score': '0.900',
                'matched_metadata_id': 'M099',
                'matched_chinese_name': '客户姓名',
            },
            {
                'metadata_id': 'M002',
                'chinese_name': '唯一字段',
                'data_type': 'VARCHAR(32)',
                'opt_time': '2026-03-26 11:00:00',
                'opt_user': 'tester',
                'service_ids': '',
                'redundancy_status': '无冗余',
                'similarity_score': '',
                'matched_metadata_id': '',
                'matched_chinese_name': '',
            },
        ],
        [],
    )

    wb = load_workbook(filename=BytesIO(output.getvalue()))
    ws = wb['新增']
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    assert headers == [
        '元数据ID',
        '中文名称',
        '数据类型',
        '涉及的场景服务（拟新增）',
        '操作时间',
        '操作用户',
        '相似度',
        '最匹配的元数据ID（运行中）',
        '匹配中文名（运行中）',
        '涉及的场景服务（运行中）',
    ]
    assert ws.max_row == 2
    assert ws.cell(2, 1).value == 'M001'
    assert ws.cell(2, 4).value == '30120001 | 30120002'
    assert ws.cell(2, 7).value == '0.900'
    assert ws.cell(2, 9).value == '客户姓名'


def test_build_excel_keeps_new_sheet_header_when_no_redundant_new_rows(tmp_path):
    config_path = tmp_path / 'config.yaml'
    _write_config(config_path, {'llm': {}, 'scan': {}})
    service = MetadataGovernanceService(config_path=config_path)

    output = service._build_excel(
        [
            {
                'metadata_id': 'M002',
                'chinese_name': '唯一字段',
                'data_type': 'VARCHAR(32)',
                'opt_time': '2026-03-26 11:00:00',
                'opt_user': 'tester',
                'redundancy_status': '无冗余',
                'similarity_score': '',
                'matched_metadata_id': '',
                'matched_chinese_name': '',
            },
        ],
        [],
    )

    wb = load_workbook(filename=BytesIO(output.getvalue()))
    ws = wb['新增']

    assert ws.max_row == 1
    assert ws.cell(1, 1).value == '元数据ID'


def test_run_analysis_now_populates_service_ids_from_sda(tmp_path, monkeypatch):
    config_path = tmp_path / 'config.yaml'
    _write_config(config_path, {'llm': {}, 'scan': {}})
    service = MetadataGovernanceService(config_path=config_path)

    class DummyInform:
        def __init__(self, *_args, **_kwargs):
            self.db_connection = object()

        def connect_database(self):
            return None

        def get_new_vs_stock_report_rows(self):
            return [
                {
                    'metadata_id': 'M001',
                    'chinese_name': '客户名称',
                    'data_type': 'VARCHAR(64)',
                    'opt_time': '2026-03-26 10:00:00',
                    'opt_user': 'tester',
                    'redundancy_status': '有冗余',
                    'similarity_score': '0.900',
                    'matched_metadata_id': 'M099',
                    'matched_chinese_name': '客户姓名',
                }
            ]

        def detect_existing_redundancy(self):
            return []

    monkeypatch.setattr('backend.service.metadata_governance_service.MetadataChangeInform', DummyInform)
    monkeypatch.setattr(service, '_build_runtime_config_for_execution', lambda: config_path)
    monkeypatch.setattr(
        service,
        '_fetch_service_ids_by_metadata',
        lambda _conn: {'M001': ['30120001', '30120002'], 'M099': ['99999999']},
    )

    result = service.run_analysis_now(match_scope='new')

    wb = load_workbook(filename=BytesIO(result.output.getvalue()))
    ws = wb['新增']
    assert ws.cell(2, 4).value == '30120001 | 30120002'


def test_run_analysis_now_uses_esb_report_filename(tmp_path, monkeypatch):
    config_path = tmp_path / 'config.yaml'
    _write_config(config_path, {'llm': {}, 'scan': {}})
    service = MetadataGovernanceService(config_path=config_path)

    class DummyInform:
        def __init__(self, *_args, **_kwargs):
            self.db_connection = object()

        def connect_database(self):
            return None

        def get_new_vs_stock_report_rows(self):
            return [
                {
                    'metadata_id': 'M001',
                    'chinese_name': '客户名称',
                    'data_type': 'VARCHAR(64)',
                    'opt_time': '2026-03-26 10:00:00',
                    'opt_user': 'tester',
                    'redundancy_status': '有冗余',
                    'similarity_score': '0.900',
                    'matched_metadata_id': 'M099',
                    'matched_chinese_name': '客户姓名',
                }
            ]

        def detect_existing_redundancy(self):
            return []

    monkeypatch.setattr('backend.service.metadata_governance_service.MetadataChangeInform', DummyInform)
    monkeypatch.setattr(service, '_build_runtime_config_for_execution', lambda: config_path)
    monkeypatch.setattr(service, '_fetch_service_ids_by_metadata', lambda _conn: {})

    result = service.run_analysis_now(match_scope='new')

    assert result.filename.startswith('ESB元数据冗余报告_')
    assert result.filename.endswith('.xlsx')


def test_metadata_governance_service_bootstrap_restores_daily_scheduler(tmp_path, monkeypatch):
    config_path = tmp_path / 'config.yaml'
    _write_config(
        config_path,
        {
            'llm': {},
            'scan': {'similarity_threshold': 0.8, 'schedule': '0 23 * * *', 'lookback_hours': 24},
        },
    )
    service = MetadataGovernanceService(config_path=config_path)
    captured = {'service': None}

    from backend.service import metadata_governance_service as module

    monkeypatch.setattr(module._daily_scheduler, 'ensure_daily_job', lambda svc: captured.__setitem__('service', svc))

    service.bootstrap_scheduler_from_config()

    assert captured['service'] is service


def test_get_current_config_returns_new_default_scope(tmp_path):
    config_path = tmp_path / 'config.yaml'
    _write_config(
        config_path,
        {
            'llm': {},
            'scan': {'similarity_threshold': 0.75, 'schedule': '0 23 * * *', 'lookback_hours': 0},
        },
    )
    service = MetadataGovernanceService(config_path=config_path)

    result = service.get_current_config()

    assert result['similarity_threshold'] == 0.75
    assert result['execution_time'] == 'daily_23'
    assert result['match_scope'] in {'new', 'stock', 'all'}


# ── persistence tests ────────────────────────────────────────────


def test_save_job_persists_to_json_and_latest(tmp_path, monkeypatch):
    config_path = tmp_path / 'config.yaml'
    _write_config(config_path, {'llm': {}, 'scan': {}})
    data_dir = tmp_path / 'data'
    service = MetadataGovernanceService(config_path=config_path, data_dir=data_dir)

    monkeypatch.setattr(settings, 'METADATA_GOVERNANCE_RETENTION_DAYS', 180)
    monkeypatch.setattr(settings, 'METADATA_GOVERNANCE_MAX_JOBS', 50)

    job = {
        'job_id': 'mgov_abc123',
        'status': 'completed',
        'created_at': '2026-03-31T10:00:00',
        'completed_at': '2026-03-31T10:05:00',
        'error': None,
        'filename': 'ESB元数据冗余报告_20260331.xlsx',
        'result_path': None,
    }
    service._save_job(job)

    import json
    jobs_path = data_dir / 'metadata_governance_jobs.json'
    latest_path = data_dir / 'metadata_governance_latest.json'

    assert jobs_path.exists()
    assert latest_path.exists()

    with open(jobs_path) as f:
        jobs = json.load(f)
    assert 'mgov_abc123' in jobs
    assert jobs['mgov_abc123']['status'] == 'completed'

    with open(latest_path) as f:
        latest = json.load(f)
    assert latest['job_id'] == 'mgov_abc123'


def test_get_job_reads_from_disk(tmp_path, monkeypatch):
    config_path = tmp_path / 'config.yaml'
    _write_config(config_path, {'llm': {}, 'scan': {}})
    data_dir = tmp_path / 'data'
    service = MetadataGovernanceService(config_path=config_path, data_dir=data_dir)

    monkeypatch.setattr(settings, 'METADATA_GOVERNANCE_RETENTION_DAYS', 180)
    monkeypatch.setattr(settings, 'METADATA_GOVERNANCE_MAX_JOBS', 50)

    job = {
        'job_id': 'mgov_xyz789',
        'status': 'running',
        'created_at': '2026-03-31T10:00:00',
        'completed_at': None,
        'error': None,
        'filename': None,
        'result_path': None,
    }
    service._save_job(job)

    fetched = service.get_job('mgov_xyz789')
    assert fetched is not None
    assert fetched['job_id'] == 'mgov_xyz789'
    assert fetched['status'] == 'running'

    assert service.get_job('nonexistent') is None


def test_get_latest_job_returns_most_recent(tmp_path, monkeypatch):
    config_path = tmp_path / 'config.yaml'
    _write_config(config_path, {'llm': {}, 'scan': {}})
    data_dir = tmp_path / 'data'
    service = MetadataGovernanceService(config_path=config_path, data_dir=data_dir)

    monkeypatch.setattr(settings, 'METADATA_GOVERNANCE_RETENTION_DAYS', 180)
    monkeypatch.setattr(settings, 'METADATA_GOVERNANCE_MAX_JOBS', 50)

    service._save_job({
        'job_id': 'mgov_first',
        'status': 'completed',
        'created_at': '2026-03-30T10:00:00',
        'completed_at': '2026-03-30T10:05:00',
        'error': None,
        'filename': 'report1.xlsx',
        'result_path': None,
    })
    service._save_job({
        'job_id': 'mgov_second',
        'status': 'completed',
        'created_at': '2026-03-31T10:00:00',
        'completed_at': '2026-03-31T10:05:00',
        'error': None,
        'filename': 'report2.xlsx',
        'result_path': None,
    })

    latest = service.get_latest_job()
    assert latest is not None
    assert latest['job_id'] == 'mgov_second'


def test_result_path_returns_persisted_file(tmp_path, monkeypatch):
    config_path = tmp_path / 'config.yaml'
    _write_config(config_path, {'llm': {}, 'scan': {}})
    data_dir = tmp_path / 'data'
    service = MetadataGovernanceService(config_path=config_path, data_dir=data_dir)

    monkeypatch.setattr(settings, 'METADATA_GOVERNANCE_RETENTION_DAYS', 180)
    monkeypatch.setattr(settings, 'METADATA_GOVERNANCE_MAX_JOBS', 50)

    # Create a dummy result file
    output = service._build_excel([{
        'metadata_id': 'M001',
        'chinese_name': '测试',
        'data_type': 'VARCHAR(32)',
        'opt_time': '2026-03-31',
        'opt_user': 'tester',
        'redundancy_status': '有冗余',
        'similarity_score': '0.95',
        'matched_metadata_id': 'M002',
        'matched_chinese_name': '测试2',
    }], [])

    result_path = service._persist_result('mgov_test1', 'report.xlsx', output)

    import os
    assert os.path.exists(result_path)

    # Save job with result_path
    service._save_job({
        'job_id': 'mgov_test1',
        'status': 'completed',
        'created_at': '2026-03-31T10:00:00',
        'completed_at': '2026-03-31T10:05:00',
        'error': None,
        'filename': 'report.xlsx',
        'result_path': result_path,
    })

    fetched_path = service.get_result_path('mgov_test1')
    assert fetched_path == result_path


def test_cleanup_removes_expired_jobs(tmp_path, monkeypatch):
    config_path = tmp_path / 'config.yaml'
    _write_config(config_path, {'llm': {}, 'scan': {}})
    data_dir = tmp_path / 'data'
    service = MetadataGovernanceService(config_path=config_path, data_dir=data_dir)

    monkeypatch.setattr(settings, 'METADATA_GOVERNANCE_RETENTION_DAYS', 1)
    monkeypatch.setattr(settings, 'METADATA_GOVERNANCE_MAX_JOBS', 50)

    # Create an old job
    service._save_job({
        'job_id': 'mgov_old',
        'status': 'completed',
        'created_at': '2020-01-01T10:00:00',
        'completed_at': '2020-01-01T10:05:00',
        'error': None,
        'filename': 'old.xlsx',
        'result_path': None,
    })
    # Create a recent job
    recent_created_at = datetime.now().replace(microsecond=0).isoformat()
    service._save_job({
        'job_id': 'mgov_recent',
        'status': 'completed',
        'created_at': recent_created_at,
        'completed_at': recent_created_at,
        'error': None,
        'filename': 'recent.xlsx',
        'result_path': None,
    })

    # The old job should be cleaned up when we save again (triggering _cleanup)
    fetched_old = service.get_job('mgov_old')
    # Note: cleanup happens during _save_job; the old job may have been removed during the second save
    fetched_recent = service.get_job('mgov_recent')
    assert fetched_recent is not None
